from concurrently import concurrently
from enum import Enum
import time
from client import *
import openpyxl
from globals.globals_season20 import *

class Rank(Enum):
    LI_M = 0
    MI_MIII = 1
    BI_DIII = 2
    R_COUNT = 3

    def to_rank(rank: int):
        if rank >= LI:
            return Rank.LI_M
        
        if rank >= MI:
            return Rank.MI_MIII
        
        return Rank.BI_DIII

class Region(Enum):
    EMEA = 0
    NA = 1
    # Add more regions here
    NONE = 2

    def to_region(region: str):
        if region == "EMEA":
            return Region.EMEA
        if region == "NA":
            return Region.NA
        # Add more regions here

        return Region.NONE
    
    def to_str(self):
        if self == Region.EMEA:
            return "EMEA"
        if self == Region.NA:
            return "NA"
        # Add more regions here

class PlayerData:
    def __init__(self):
        self.tag = ""
        self.l_game_checked = None

    def print(self):
        print("Player_data : [tag : {}, l_game_checked : {}]".format(self.tag, self.l_game_checked))

    def str(self):
        if self.l_game_checked is None:
            return self.tag + "\n"
        else:
            return self.tag + " " + self.l_game_checked + "\n"
    
class BattleData:
    def __init__(self, map: str):
        # todo : move self.map initialization is set_battle_data func.
        self.map = map
        self.rank = Rank.R_COUNT
        self.winner_brawlers = []
        self.loser_brawlers = []

    def set_battle_data(self, battle: dict, winner_side: int, loser_side: int):
        winner_team = battle["teams"][winner_side] 
        loser_team = battle["teams"][loser_side]

        rank = 1

        for player in winner_team:
            self.winner_brawlers.append(player["brawler"]["name"])
            rank = max(rank, player["brawler"]["trophies"])

        for player in loser_team:
            self.loser_brawlers.append(player["brawler"]["name"])
            rank = max(rank, player["brawler"]["trophies"])

        self.rank = Rank.to_rank(rank)

    def print(self):
        print("Battle_data : [map : {0}, rank : {1}, winner_brawlers : {2}, loser_brawlers : {3}]".format(self.map, self.rank, self.winner_brawlers, self.loser_brawlers))

def get_teams_status(tag: str, battle: dict, match_status: bool):
    for player in battle["teams"][0]:
        if tag == player["tag"]:
            if match_status == True:
                return (0, 1)
            else:
                return (1, 0)
        
    if match_status == True:
        return (1, 0)
    else:
        return (0, 1)

def set_player_fields(fields: list):
    player = PlayerData()
    fields_len = len(fields)
    if fields_len > 0:
        player.tag = fields[0]
    if fields_len > 1:
        player.l_game_checked = fields[1]

    return player

def tag_not_in_list(tag: str, new_players: list):        
    for player in new_players:
        if tag == player.tag:
            return False

    return True

def add_new_player(new_players: list, tag: str):
    new_player = PlayerData()
    new_player.tag = tag
    new_players.append(new_player)

def add_players(new_players: list, battle: dict):
    teams = battle["teams"]
    for team in teams:
        for player in team:
            if tag_not_in_list(player["tag"], new_players) == True:
                add_new_player(new_players, player["tag"])

def read_player_data():
    players = []
    f = open("player_tags.txt", "r")
    for line in f:
        fields = line[0:len(line) - 1].split()
        players.append(set_player_fields(fields))

    f.close()

    print("Reading player data done. {0} players to be checked for data collection.".format(len(players)))

    return players

def update_player_data(players: list):
    f = open("player_tags.txt", "w")
    for player in players:
        f.write(player.str())

    f.close()

    print("Updating player data done.")

def get_battles(player: PlayerData) -> list:
    battles_data = []
    new_players = []

    battlelog = Client.get_player_battlelog(player.tag)

    if battlelog is None:
        return (battles_data, new_players)
    
    new_battles = 0
    battles_dict = battlelog["items"]
    for battle_dict in reversed(battles_dict):
        battle = battle_dict["battle"]

        if battle.get("type") is None:
            continue
            
        if battle.get("type") != "soloRanked":
            continue
            
        if player.l_game_checked is not None and battle_dict["battleTime"] <= player.l_game_checked:
            continue

        if MAPS.get(battle_dict["event"]["map"]) is None:
            continue
            
        if battle.get("starPlayer") is not None:
            add_players(new_players, battle)
            new_battles += 1
            player.l_game_checked = battle_dict["battleTime"]
            match_status = True
            if battle["result"] == "defeat":
                match_status = False

            battle_data = BattleData(battle_dict["event"]["map"])
            (winner_side, loser_side) = get_teams_status(player.tag, battle, match_status)
            battle_data.set_battle_data(battle, winner_side, loser_side)
            battles_data.append(battle_data)

    return (battles_data, new_players)

def collect_pl_data(players: list) -> list:
    battles_data: list = []
    new_players_data: list = []

    for (_, (player_battles, new_players)) in concurrently(handler=get_battles, inputs=players, max_concurrency=MAX_CONCURRENCY):
        battles_data.extend(player_battles)
        for new_player in new_players:
            if len(players) + len(new_players_data) < MAX_NR_PLAYERS \
                and tag_not_in_list(new_player.tag, players) == True \
                and tag_not_in_list(new_player.tag, new_players_data) == True:
                new_players_data.append(new_player)

    players.extend(new_players_data)

    print("Finished to retrieve battlelogs.\nTotal new battles found : {}.\nTotal new players added : {}".format(len(battles_data), len(new_players_data)))

    return battles_data

def save_pl_data(battles_data: list):
    wb_obj = openpyxl.load_workbook(FILE_PATH)
    new_battles = 0

    for battle_data in battles_data:
        sheet = wb_obj[wb_obj.sheetnames[battle_data.rank.value]]

        (picks_col, wins_col) = MAPS.get(battle_data.map)
        if picks_col is None and wins_col is None:
            print("Error. Couldn't retrieve collums for map {0}.".format(battle_data.map))
            continue

        new_battles += 1
        
        for brawler in battle_data.winner_brawlers:
            brawler_row = BRAWLERS.get(brawler)
            if brawler_row is None:
                print("Error. Couldn't retrieve row for brawler {0}.".format(brawler))
                continue

            picks_cell = sheet.cell(row=brawler_row, column=picks_col)
            wins_cell = sheet.cell(row=brawler_row, column=wins_col)

            picks_cell.value = int(picks_cell.value) + 1
            wins_cell.value = int(wins_cell.value) + 1

        for brawler in battle_data.loser_brawlers:
            brawler_row = BRAWLERS.get(brawler)
            if brawler_row is None:
                print("Error. Couldn't retrieve row for brawler {0}.".format(brawler))
                continue

            picks_cell = sheet.cell(row=brawler_row, column=picks_col)

            picks_cell.value = int(picks_cell.value) + 1

    wb_obj.save(FILE_PATH)

    print("Saving player data done. {0} new battles registered.".format(new_battles))

players = read_player_data()

try:
    while True:
        start = time.time()
        battles_data = collect_pl_data(players)
        if len(battles_data) != 0:
            update_player_data(players)
            save_pl_data(battles_data)
        else:
            print("Updating player data skipped. No new battes.")
            print("Saving player data skipped. No new battes.")
        end = time.time()
        print("Time taken : {} min, {} seconds.".format((end - start) // 60, round((end - start) % 60, 2)))
except KeyboardInterrupt:
    pass
