from enum import Enum
import openpyxl
from globals.globals_season20 import *
import shutil
import os
# from git import Repo
import tempfile
import pygame
from pygame.locals import *

pygame.init()
frame_rate = pygame.time.Clock()

PATH = os.path.abspath('.') + '/'

WHITE = (255,255,255)
RED = (255,0,0)
DARKR = (125, 0, 0)
GREEN = (0,255,0)
BLACK = (0,0,0)
YELLOW = (255,255,0)
ORANGE = (255,165,0)
DARKG = (0,100,0)
BROWN = (165,42,42)
GOLD = (255,215,0)
SILVER = (192,192,192)
BRONZE = (205,127,50)
LIGHTB = (37, 145, 246)
PURPLE = (128, 0, 128)

#globals
WIDTH = 1024
HEIGHT = 768

info = pygame.display.Info()
APP_WIDTH = info.current_w - 10
APP_HEIGHT = info.current_h - 100

SCALE_X = APP_WIDTH / WIDTH
SCALE_Y = APP_HEIGHT / HEIGHT

def get_color_by_percentage(percentage: str):
    percent = float(percentage)

    if percent > 80.0:
        return PURPLE

    if percent > 60.0:
        return GREEN

    if percent < 10.0:
        return RED

    if percent < 30.0:
        return YELLOW

    return WHITE

class State(Enum):
    HOME = 0
    GAMEMODES_HOME = 1
    MAPS_HOME = 2
    RANKINGS = 3

class Image:
    def __init__(self, path: str, pos: tuple, name: str, name_pos: tuple):
        self.path = path
        self.pos = (pos[0] * SCALE_X, pos[1] * SCALE_Y)
        self.name = name
        self.name_pos = (name_pos[0] * SCALE_X, name_pos[1] * SCALE_Y)

    def draw(self, window):
        img = pygame.image.load(PATH + self.path)
        img = pygame.transform.scale(img, (370 * SCALE_X, 520 * SCALE_Y))
        window.blit(img, self.pos)

        myfont = pygame.font.SysFont("Comic Sans MS", 50)
        label = myfont.render(self.name , 1, (0,0,0))
        window.blit(label, (self.name_pos[0], self.name_pos[1]))
        pass

class Brawler:
    def __init__(self, prof_icon: Image, winrate: str, playrate: str):
        self.prof_icon = prof_icon
        self.winrate = winrate
        self.playrate = playrate

    def __lt__(self, other):
        return float(self.winrate) * 0.3 + float(self.playrate) * 0.7 > float(other.winrate) * 0.3 + float(other.playrate) * 0.7

    def draw(self, window):
        img = pygame.image.load(PATH + self.prof_icon.path)
        img = pygame.transform.scale(img, (40 * SCALE_X, 40 * SCALE_Y))
        window.blit(img, self.prof_icon.pos)

        myfont = pygame.font.SysFont("Comic Sans MS", 40)
        label = myfont.render(self.prof_icon.name, 1, (0,0,0))
        window.blit(label, (self.prof_icon.name_pos[0], self.prof_icon.name_pos[1]))

        label = myfont.render(self.winrate + "%", 1, get_color_by_percentage(self.winrate))
        window.blit(label, (740 * SCALE_X, self.prof_icon.name_pos[1]))

        label = myfont.render(self.playrate + "%", 1, get_color_by_percentage(self.playrate))
        window.blit(label, (890 * SCALE_X, self.prof_icon.name_pos[1]))

    def print(self):
        print("{} {} {}".format(self.prof_icon.name, self.winrate, self.playrate))

class Button:
    def __init__(self, name: str, pos: tuple, dimensions: tuple):
        self.name = name
        self.pos = (pos[0] * SCALE_X, pos[1] * SCALE_Y)
        self.dim = (dimensions[0] * SCALE_X, dimensions[1] * SCALE_Y)
        self.text_pos = (self.pos[0] + self.dim[0] // 8, self.pos[1] + self.dim[1] // 4)
        self.hovered = False

    def draw(self, window):
        pygame.draw.rect(window, RED, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]))
        if self.hovered == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]), 2)

        myfont = pygame.font.SysFont("Comic Sans MS", 50)
        label = myfont.render(self.name , 1, (0,0,0))
        window.blit(label, (self.text_pos[0], self.text_pos[1]))

    def is_pressed(self, x, y):
        if (x > self.pos[0] and x < self.pos[0] + self.dim[0]) and \
            (y > self.pos[1] and y < self.pos[1] + self.dim[1]):
            return True
        
        return False
    
    def is_hovered(self, x, y):
        if (x > self.pos[0] and x < self.pos[0] + self.dim[0]) and \
            (y > self.pos[1] and y < self.pos[1] + self.dim[1]):
            return True
        
        return False    

    def on_press(self, curr_state: State):
        return State(curr_state.value + 1)

class ReturnButton(Button):
    def __init__(self, name: str, pos: tuple, dimensions: tuple):
        super().__init__(name, pos, dimensions)
        self.text_pos = (self.pos[0] + int(self.dim[0] / 3.5), self.pos[1] + self.dim[1] / 4)

    def draw(self, window):
        pygame.draw.rect(window, RED, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]))
        if self.hovered == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]), 2)

        myfont = pygame.font.SysFont("Comic Sans MS", 45)
        label = myfont.render(self.name , 1, (0,0,0))
        window.blit(label, (self.text_pos[0], self.text_pos[1]))

    def on_press(self, curr_state):
        return State(curr_state.value - 1)

class StatsButton(Button):
    def __init__(self, name: str, pos: tuple, dimensions: tuple):
        super().__init__(name, pos, dimensions)
        self.text_pos = (self.pos[0] + int(self.dim[0] / 3.5), self.pos[1] + self.dim[1] / 3)

    def on_press(self, _):
        return State.GAMEMODES_HOME

class ExitButton(Button):
    def __init__(self, name: str, pos: tuple, dimensions: tuple):
        super().__init__(name, pos, dimensions)
        self.text_pos = (self.pos[0] + self.dim[0] / 3, self.pos[1] + self.dim[1] / 3)

    def on_press(self, _):
        # os.remove(PATH + FILE_QUERY_PATH)
        exit(0)

class GameModeButton(Button):
    def draw(self, window):
        img = pygame.image.load(PATH + "pictures/gamemodes/{}.png".format(self.name))
        img = pygame.transform.scale(img, (self.dim[0], self.dim[1]))
        window.blit(img, self.pos)

        if self.hovered == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]), 2)
    
class MapButton(Button):
    def __init__(self, name: str, img_path: str, pos: tuple, dimensions: tuple):
        super().__init__(name, pos, dimensions)
        self.img_path = PATH + img_path

    def draw(self, window):
        img = pygame.image.load(self.img_path)
        img = pygame.transform.scale(img, (self.dim[0], self.dim[1]))
        window.blit(img, self.pos)

        if self.hovered == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]), 2)

class RankButton(Button):
    def __init__(self, name: str, pos: tuple, dimensions: tuple, id: int):
        super().__init__(name, pos, dimensions)
        self.text_pos = (self.pos[0] + int(self.dim[0] / 3.5), self.pos[1] + self.dim[1] / 3)
        self.selected = False
        self.id = id
        if id == 0:
            self.selected = True

    def draw(self, window):
        if self.selected == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]))
        else:
            pygame.draw.rect(window, RED, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]))
        if self.hovered == True:
            pygame.draw.rect(window, DARKR, (self.pos[0], self.pos[1], self.dim[0], self.dim[1]), 2)

        myfont = pygame.font.SysFont("Comic Sans MS", 25)
        label = myfont.render(self.name , 1, (0,0,0))
        window.blit(label, (self.text_pos[0], self.text_pos[1]))

    def on_press(self, _):
        return self.id

class Screen:
    def __init__(self):
        self.background = PATH + "pictures/mini_brawl_stats.jpg"
        self.buttons = []

    def draw(self, window):
        bg_img = pygame.image.load(self.background)
        bg_img = pygame.transform.scale(bg_img,(WIDTH * SCALE_X, HEIGHT * SCALE_Y))
        window.blit(bg_img, (0, 0))

        for button in self.buttons:
            button.draw(window)

    def input(self, x, y):
        for button in self.buttons:
            if button.is_pressed(x, y) == True:
                return button

        return None
    
    def mouse_move(self, x, y):
        for button in self.buttons:
            if button.is_hovered(x, y) == True:
                button.hovered = True
            else:
                button.hovered = False

class HomeScreen(Screen):
    def __init__(self):
        super().__init__()
        self.buttons = [
            StatsButton("Stats", (350, 150), (200, 100)), 
            ExitButton("Exit", (350, 300), (200, 100))
        ]

class StatsScreen(Screen):
    def __init__(self):
        super().__init__()
        f = open(PATH + "gamemodes/{}/curr_rotation.txt".format(SEASON))
        self.buttons = [
            GameModeButton(f.readline().strip(), (15, 80), (270, 90)), 
            GameModeButton(f.readline().strip(), (375, 80), (270, 90)),
            GameModeButton(f.readline().strip(), (735, 80), (270, 90)), 
            GameModeButton(f.readline().strip(), (15, 200), (270, 90)),
            GameModeButton(f.readline().strip(), (375, 200), (270, 90)), 
            GameModeButton(f.readline().strip(), (735, 200), (270, 90)),
            ReturnButton("Back", (0, 0), (150, 55))
        ]

class GameModeScreen(Screen):
    def __init__(self):
        super().__init__()

    def update_layout(self, file_name):
        img_path = "pictures/map_icons/{}/".format(file_name)
        f = open(PATH + "maps/{}/{}_maps.txt".format(SEASON, file_name))
        g_mode1 = f.readline().strip()
        g_mode2 = f.readline().strip()
        g_mode3 = f.readline().strip()
        self.buttons = [
            MapButton(g_mode1, img_path + g_mode1 + "_icon.png", (15, 100), (300, 150)), 
            MapButton(g_mode2, img_path + g_mode2 + "_icon.png", (360, 100), (300, 150)),
            MapButton(g_mode3, img_path + g_mode3 + "_icon.png", (705, 100), (300, 150)),
            ReturnButton("Back", (0, 0), (150, 55))
        ]

class MapScreen(Screen):
    def __init__(self):
        super().__init__()
        self.map = None
        self.curr_rank = 0
        self.top_brawlers = [[], [], []]
        self.buttons = [
            RankButton("LI - M", (40, 90), (120, 75), 0),
            RankButton("MI - MIII", (175, 90), (120, 75), 1),
            RankButton("BI - DIII", (310, 90), (120, 75), 2),
            ReturnButton("Back", (0, 0), (150, 55))
        ]

    def draw(self, window):
        bg_img = pygame.image.load(self.background)
        bg_img = pygame.transform.scale(bg_img,(WIDTH * SCALE_X, HEIGHT * SCALE_Y))
        window.blit(bg_img, (0, 0))

        pygame.draw.rect(window, LIGHTB, (30 * SCALE_X, 60 * SCALE_Y, (WIDTH - 60) * SCALE_X, (HEIGHT - 70) * SCALE_Y))

        for button in self.buttons:
            button.draw(window)

        self.map.draw(window)

        myfont = pygame.font.SysFont("Comic Sans MS", 50)
        label = myfont.render("Nr." , 1, (0,0,0))
        window.blit(label, (450 * SCALE_X, 60 * SCALE_Y))

        label = myfont.render("Brawler" , 1, (0,0,0))
        window.blit(label, (525 * SCALE_X, 60 * SCALE_Y))

        label = myfont.render("Wr" , 1, (0,0,0))
        window.blit(label, (750 * SCALE_X, 60 * SCALE_Y))

        label = myfont.render("Pr" , 1, (0,0,0))
        window.blit(label, (910 * SCALE_X, 60 * SCALE_Y))

        label = myfont.render("1" , 1, (0,0,0))
        pygame.draw.rect(window, GOLD, ((450 + 8) * SCALE_X, (55 + 44 * 1) * SCALE_Y, 30, 30))
        window.blit(label, ((450 + 13) * SCALE_X, (55 + 44 * 1) * SCALE_Y))

        label = myfont.render("2" , 1, (0,0,0))
        pygame.draw.rect(window, SILVER, ((450 + 8) * SCALE_X, (55 + 44 * 2) * SCALE_Y, 30, 30))
        window.blit(label, ((450 + 13) * SCALE_X, (55 + 44 * 2) * SCALE_Y))

        label = myfont.render("3" , 1, (0,0,0))
        pygame.draw.rect(window, BRONZE, ((450 + 8) * SCALE_X, (55 + 44 * 3) * SCALE_Y, 30, 30))
        window.blit(label, ((450 + 13) * SCALE_X, (55 + 44 * 3) * SCALE_Y))

        for i in range(4, 16, 1):
            label = myfont.render(str(i) , 1, (0,0,0))
            if i < 10:
                window.blit(label, ((450 + 13) * SCALE_X, (55 + 44 * i) * SCALE_Y))
            else:
                window.blit(label, (450 * SCALE_X, (55 + 44 * i) * SCALE_Y))

        for brawler in self.top_brawlers[self.curr_rank]:
            brawler.draw(window)

    def update_layout(self, file_name, wb_obj: openpyxl.Workbook):
        self.map = None
        self.curr_rank = 0
        self.top_brawlers = [[], [], []]
        for i in range(0, len(self.buttons) - 1):
            if i == self.curr_rank:
                self.buttons[i].selected = True
            else:
                self.buttons[i].selected = False

        self.map = Image("pictures/maps/" + file_name + ".png", (50, 230), file_name, (50, 185))
        for rank in range(0, 3):
            sheet = wb_obj[wb_obj.sheetnames[rank]]
            total_games = 0
            brawlers = []
            for i in range(5, len(BRAWLERS)+ 5):
                cell_obj = sheet.cell(row = i, column = MAPS[self.map.name][0])
                total_games += int(cell_obj.value)

            total_games //= 6
            for i in range(5, len(BRAWLERS)+ 5):
                
                winrate = ""
                playrate = ""
                br_name = sheet.cell(row = i, column = 1)
                br_picks = sheet.cell(row = i, column = MAPS[self.map.name][0]) 
                br_wins = sheet.cell(row = i, column = MAPS[self.map.name][1])

                if total_games == 0:
                    playrate = "0.0"
                else:
                    playrate = str(round(int(br_picks.value) * 100 / total_games, 2))

                if int(br_picks.value) == 0:
                    winrate = "0.0"
                else:
                    winrate = str(round(int(br_wins.value) * 100 / int((br_picks).value), 2))

                brawlers.append(Brawler(Image(
                                            "pictures/brawlers_pictures/" + br_name.value + ".png", 
                                            (0, 0),
                                            br_name.value, 
                                            (0, 0)
                                        ), 
                                    winrate, 
                                    playrate
                                )
                )

            brawlers.sort()

            for i in range(0, 15):
                self.top_brawlers[rank].append(Brawler(Image(brawlers[i].prof_icon.path, (510, 55 + 44 * (i + 1)), brawlers[i].prof_icon.name, (560, 55 + 44 * (i + 1))), brawlers[i].winrate, brawlers[i].playrate))

class App:
    def __init__(self):
        os.environ['SDL_VIDEO_CENTERED'] = '1'
        self.window = pygame.display.set_mode((APP_WIDTH, APP_HEIGHT), pygame.SCALED | pygame.FULLSCREEN)
        pygame.display.set_caption('Power League Stats')

        self.wb_obj = openpyxl.load_workbook(PATH + FILE_QUERY_PATH)

        self.running = True
        self.curr_state = State.HOME
        self.pages = [HomeScreen(), StatsScreen(), GameModeScreen(), MapScreen()]

    def update(self):
        pass

    def input(self):
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                self.running = False

            if event.type == pygame.MOUSEBUTTONDOWN:
                mouse_presses = pygame.mouse.get_pressed()
                if mouse_presses[0]:
                    (pos_x, pos_y) = pygame.mouse.get_pos()
                    button_pressed = self.pages[self.curr_state.value].input(pos_x, pos_y)
                    if button_pressed is not None:
                        if type(button_pressed) is not RankButton:
                            self.curr_state = button_pressed.on_press(self.curr_state)
                            if button_pressed.name != "Back":
                                if self.curr_state == State.MAPS_HOME:
                                    self.pages[self.curr_state.value].update_layout(button_pressed.name)

                                if self.curr_state == State.RANKINGS:
                                    self.pages[self.curr_state.value].update_layout(button_pressed.name, self.wb_obj)
                        else:
                            aux_rank = self.pages[self.curr_state.value].curr_rank
                            self.pages[self.curr_state.value].buttons[aux_rank].selected = False 

                            self.pages[self.curr_state.value].curr_rank = button_pressed.on_press(self.curr_state)

                            aux_rank = self.pages[self.curr_state.value].curr_rank
                            self.pages[self.curr_state.value].buttons[aux_rank].selected = True

            if pygame.mouse.get_focused() == True:
                (pos_x, pos_y) = pygame.mouse.get_pos()
                self.pages[self.curr_state.value].mouse_move(pos_x, pos_y)


    def draw(self):
        self.pages[self.curr_state.value].draw(self.window)
        pygame.display.update()
        frame_rate.tick(60)

# # Create temporary dir
# t = tempfile.mkdtemp()
# # Clone into temporary dir
# Repo.clone_from('https://github.com/Stefan19K/BS_Power_League_Tracker', t, branch='main', depth=1)
# # Copy desired file from temporary dir
# shutil.move(os.path.join(t, FILE_PATH), FILE_QUERY_PATH)
# # Remove temporary dir
# shutil.rmtree(t)

app = App()

while app.running:
    app.draw()
    app.update()
    app.input()

# os.remove(PATH + FILE_QUERY_PATH)
